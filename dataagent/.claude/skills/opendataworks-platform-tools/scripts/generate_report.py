from __future__ import annotations

import argparse
import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from _opendataworks_runtime import error_payload, load_json_input, print_json


def _resolve_path(raw: str) -> Path:
    path = Path(str(raw or "").strip()).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    return path


def style_excel_sheet(excel_path: Path):
    """Use openpyxl to apply a premium design to the generated Excel report."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Premium Navy/Blue palette
        header_fill = PatternFill(
            start_color="1B365D", end_color="1B365D", fill_type="solid"
        )
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        zebra_fill = PatternFill(
            start_color="F2F4F8", end_color="F2F4F8", fill_type="solid"
        )

        thin_side = Side(border_style="thin", color="D3D3D3")
        border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # Set header styles
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border

        # Set data row styles
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
            start=2,
        ):
            use_zebra = row_idx % 2 == 0
            for cell in row:
                cell.font = data_font
                cell.border = border
                if use_zebra:
                    cell.fill = zebra_fill
                # Alignment based on value type
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left

        # Adjust column widths dynamically
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(excel_path)
    except Exception:
        # If styling fails, we still keep the original excel file
        pass


def generate_html_report(data_rows: list[dict[str, Any]], title: str, output_path: Path):
    """Generate a premium responsive HTML report with modern CSS styling."""
    if not data_rows:
        html = "<html><body><p>No data available</p></body></html>"
        output_path.write_text(html, encoding="utf-8")
        return

    columns = list(data_rows[0].keys())

    # Build table rows
    table_rows = []
    for idx, row in enumerate(data_rows):
        row_class = "even" if idx % 2 == 0 else "odd"
        cells = []
        for col in columns:
            val = row.get(col, "")
            is_num = False
            try:
                if val is not None:
                    float(str(val))
                    is_num = True
            except ValueError:
                pass
            align_class = " align-right" if is_num else ""
            cells.append(f"<td class='{align_class}'>{val}</td>")
        table_rows.append(f"<tr class='{row_class}'>{''.join(cells)}</tr>")

    headers = "".join(f"<th>{col}</th>" for col in columns)

    # CSS styles for a premium design
    css = """
    body {
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
        background-color: #f8f9fa;
        color: #333333;
        margin: 0;
        padding: 40px 20px;
    }
    .container {
        max-width: 1000px;
        margin: 0 auto;
        background-color: #ffffff;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.05);
        padding: 30px;
    }
    h1 {
        color: #1b365d;
        font-size: 24px;
        margin-top: 0;
        margin-bottom: 20px;
        border-bottom: 2px solid #f2f4f8;
        padding-bottom: 10px;
    }
    .meta {
        font-size: 13px;
        color: #666666;
        margin-bottom: 20px;
    }
    table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 10px;
        border-radius: 8px;
        overflow: hidden;
    }
    th {
        background-color: #1b365d;
        color: #ffffff;
        font-weight: 600;
        font-size: 14px;
        padding: 12px 15px;
        text-align: left;
    }
    td {
        padding: 10px 15px;
        border-bottom: 1px solid #e9ecef;
        font-size: 13.5px;
    }
    .even {
        background-color: #f8f9fa;
    }
    .odd {
        background-color: #ffffff;
    }
    tr:hover {
        background-color: #f1f3f5;
    }
    .align-right {
        text-align: right;
    }
    """

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{title}</title>
    <style>{css}</style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        <div class="meta">生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        <table>
            <thead>
                <tr>{headers}</tr>
            </thead>
            <tbody>
                {"".join(table_rows)}
            </tbody>
        </table>
    </div>
</body>
</html>
"""
    output_path.write_text(html, encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel (.xlsx) or HTML reports from CSV/JSON input"
    )
    parser.add_argument("--input", required=True, help="Input CSV file path or raw JSON array")
    parser.add_argument("--output", required=True, help="Output file path (e.g. output/report.xlsx or output/report.html)")
    parser.add_argument("--title", default="数据分析报告", help="Report Title")
    args = parser.parse_args()

    input_val = str(args.input or "").strip()
    output_path = _resolve_path(args.output)
    title = str(args.title or "").strip()

    try:
        rows: list[dict[str, Any]] = []
        if input_val.endswith(".csv") or os.path.exists(input_val):
            csv_path = _resolve_path(input_val)
            if not csv_path.exists():
                raise FileNotFoundError(f"输入文件不存在: {csv_path}")
            
            with open(csv_path, "r", encoding="utf-8-sig") as handle:
                reader = csv.DictReader(handle)
                rows = [dict(row) for row in reader]
        else:
            loaded = load_json_input(raw=input_val)
            if isinstance(loaded, list):
                rows = [dict(row) for row in loaded if isinstance(row, dict)]
            elif isinstance(loaded, dict) and "rows" in loaded:
                rows = [dict(row) for row in loaded["rows"] if isinstance(row, dict)]
            else:
                raise ValueError("不支持的 JSON 格式，必须是 JSON 数组或包含 rows 字段的 JSON 对象")

        if not rows:
            raise ValueError("没有可用于生成报告的数据")

        suffix = output_path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            df = pd.DataFrame(rows)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_excel(output_path, index=False, engine="openpyxl")
            style_excel_sheet(output_path)
            format_name = "Excel 报表"
        elif suffix in (".html", ".htm"):
            output_path.parent.mkdir(parents=True, exist_ok=True)
            generate_html_report(rows, title, output_path)
            format_name = "HTML 报告"
        else:
            raise ValueError(f"不支持的输出格式: {suffix}。请提供以 .xlsx 或 .html 结尾的输出路径。")

        file_size = output_path.stat().st_size
        print_json(
            {
                "kind": "report_generation",
                "tool_label": "报告生成",
                "status": "success",
                "title": title,
                "input": input_val,
                "output_path": str(output_path),
                "format": format_name,
                "file_size": file_size,
                "row_count": len(rows),
                "summary": f"已成功生成 {format_name}（{len(rows)} 行数据，大小 {file_size} 字节）并写入至 {output_path}",
                "error": None,
            }
        )

    except Exception as exc:
        print_json(
            error_payload(
                "report_generation",
                str(exc),
                tool_label="报告生成",
                status="failed",
                title=title,
                input=input_val,
                output_path=str(output_path),
                summary="生成报告失败，请检查输入数据或输出路径格式后重试。",
            )
        )


if __name__ == "__main__":
    main()
